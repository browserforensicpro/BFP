"""
reportbuilder.py - Generate forensic reports in PDF, HTML, CSV, Excel, XML.
"""

import os
import csv
import json
import logging
import datetime
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


def export_csv(data: list, filepath: str, columns: list = None) -> bool:
    """Export data list to CSV."""
    try:
        if not data:
            return False
        if columns is None:
            columns = list(data[0].keys()) if data else []
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(data)
        logger.info(f"CSV exported: {filepath}")
        return True
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return False


def export_excel(data: list, filepath: str, sheet_name: str = "Forensic Data") -> bool:
    """Export data to Excel XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(filepath)
            return True

        columns = list(data[0].keys())
        # Header row
        header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
        header_font = Font(color="00bfff", bold=True)
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(15, len(col) + 4)

        # Data rows
        for row_idx, row in enumerate(data, 2):
            fill_color = "1e293b" if row_idx % 2 == 0 else "162032"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            white_font = Font(color="e2e8f0")
            for col_idx, col in enumerate(columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(row.get(col, "")))
                cell.fill = fill
                cell.font = white_font

        wb.save(filepath)
        logger.info(f"Excel exported: {filepath}")
        return True
    except ImportError:
        logger.warning("openpyxl not installed. Install: pip install openpyxl")
        return False
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return False


def export_xml(data: list, filepath: str, root_tag: str = "ForensicData",
               item_tag: str = "Record") -> bool:
    """Export data to XML."""
    try:
        root = ET.Element(root_tag)
        root.set("generated", datetime.datetime.now().isoformat())
        root.set("count", str(len(data)))
        for row in data:
            item = ET.SubElement(root, item_tag)
            for key, val in row.items():
                child = ET.SubElement(item, key.replace(" ", "_").replace("/", "_"))
                child.text = str(val or "")
        tree = ET.ElementTree(root)
        ET.indent(tree, space="  ")
        tree.write(filepath, encoding="utf-8", xml_declaration=True)
        logger.info(f"XML exported: {filepath}")
        return True
    except Exception as e:
        logger.error(f"XML export error: {e}")
        return False


def export_html(data: list, filepath: str, title: str = "Forensic Report",
                metadata: dict = None) -> bool:
    """Generate a styled HTML forensic report."""
    try:
        if metadata is None:
            metadata = {}
        columns = list(data[0].keys()) if data else []
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        rows_html = ""
        for i, row in enumerate(data):
            bg = "#1e293b" if i % 2 == 0 else "#162032"
            cells = "".join(f"<td>{str(row.get(c, ''))}</td>" for c in columns)
            rows_html += f'<tr style="background:{bg}">{cells}</tr>\n'

        headers_html = "".join(f"<th>{c}</th>" for c in columns)
        meta_html = "".join(
            f"<p><strong>{k}:</strong> {v}</p>"
            for k, v in metadata.items()
        )

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{title}</title>
<style>
  body {{ background:#0f172a; color:#e2e8f0; font-family:'Segoe UI',sans-serif; margin:0; padding:20px; }}
  h1 {{ color:#00bfff; border-bottom:2px solid #00bfff; padding-bottom:10px; }}
  .meta {{ background:#1e293b; padding:15px; border-left:4px solid #00bfff; margin-bottom:20px; border-radius:4px; }}
  table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  th {{ background:#0f3460; color:#00bfff; padding:8px; text-align:left; position:sticky; top:0; }}
  td {{ padding:6px 8px; border-bottom:1px solid #334155; }}
  tr:hover td {{ background:#0f3460 !important; }}
  .footer {{ color:#64748b; font-size:11px; margin-top:20px; text-align:center; }}
</style>
</head>
<body>
<h1>🔍 {title}</h1>
<div class="meta">
  <p><strong>Generated:</strong> {now}</p>
  <p><strong>Total Records:</strong> {len(data)}</p>
  {meta_html}
</div>
<div style="overflow-x:auto">
<table>
<thead><tr>{headers_html}</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</div>
<div class="footer">Browser Forensics Pro — Digital Forensics Report</div>
</body>
</html>"""
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html)
        logger.info(f"HTML report exported: {filepath}")
        return True
    except Exception as e:
        logger.error(f"HTML export error: {e}")
        return False


def export_pdf(data: list, filepath: str, title: str = "Forensic Report",
               metadata: dict = None) -> bool:
    """Generate PDF report using reportlab."""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer)
        from reportlab.lib.units import inch

        if metadata is None:
            metadata = {}

        doc = SimpleDocTemplate(
            filepath,
            pagesize=landscape(A4),
            leftMargin=0.5*inch, rightMargin=0.5*inch,
            topMargin=0.5*inch, bottomMargin=0.5*inch,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "title_style",
            fontSize=16, textColor=colors.HexColor("#00bfff"),
            spaceAfter=12, fontName="Helvetica-Bold",
        )
        story.append(Paragraph(f"🔍 {title}", title_style))

        # Metadata
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        meta_style = ParagraphStyle(
            "meta_style", fontSize=9,
            textColor=colors.HexColor("#94a3b8"), spaceAfter=4,
        )
        story.append(Paragraph(f"Generated: {now} | Records: {len(data)}", meta_style))
        for k, v in metadata.items():
            story.append(Paragraph(f"{k}: {v}", meta_style))
        story.append(Spacer(1, 12))

        if not data:
            story.append(Paragraph("No data to display.", meta_style))
            doc.build(story)
            return True

        columns = list(data[0].keys())
        max_cols = min(len(columns), 10)  # Limit columns for readability
        columns = columns[:max_cols]

        # Table data
        header_row = [Paragraph(c, ParagraphStyle(
            "hdr", fontSize=7, textColor=colors.HexColor("#00bfff"),
            fontName="Helvetica-Bold")) for c in columns]
        table_data = [header_row]

        cell_style = ParagraphStyle("cell", fontSize=6.5,
                                     textColor=colors.HexColor("#e2e8f0"))
        for row in data[:2000]:  # PDF limit
            table_row = []
            for col in columns:
                val = str(row.get(col, ""))[:80]
                table_row.append(Paragraph(val, cell_style))
            table_data.append(table_row)

        col_width = (landscape(A4)[0] - inch) / max_cols
        tbl = Table(table_data, colWidths=[col_width] * max_cols, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#1e293b")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#1e293b"), colors.HexColor("#162032")]),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#334155")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)

        doc.build(story)
        logger.info(f"PDF exported: {filepath}")
        return True
    except ImportError:
        logger.warning("reportlab not installed. Install: pip install reportlab")
        return False
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return False
